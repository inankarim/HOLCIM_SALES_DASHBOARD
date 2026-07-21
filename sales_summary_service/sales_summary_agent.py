#!/usr/bin/env python3
"""
sales_summary_agent.py
=======================

Consolidates two daily Excel exports (Supercrete PLC/PLC+/Powercrete
distribution, and PCC+OPC/HWP/HCG distribution) into a single formatted
"Sales_Summary.xlsx" workbook, matched by SAP ID.

Design goals (per spec):
  - No hardcoded filenames. Files are identified by their column fingerprints.
  - Column names are normalized before matching, so header whitespace/line
    breaks/casing/reordering never breaks detection.
  - Extra/unknown columns are ignored. Missing *required* columns raise a
    clear, itemized error instead of failing silently or crashing.
  - Vectorized pandas operations only (no per-row Python loops) so it stays
    fast on 100k+ row files.
  - openpyxl formatting: two-row grouped headers, frozen panes, borders,
    zebra striping, number/percent formats, and conditional formatting on
    every Achievement % column.

Usage:
    python sales_summary_agent.py <file_a.xlsx> <file_b.xlsx> [-o Sales_Summary.xlsx]

The two input files can be passed in either order — the agent detects
which one is the "PLC/Powercrete" file and which is the "PCC+OPC/HWP/HCG"
file by inspecting their columns, not their filenames or position.
"""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass, field
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# ---------------------------------------------------------------------------
# 1. Configuration: customer-info columns + product column groups
# ---------------------------------------------------------------------------

# Customer/rep info columns that should appear exactly once in the output.
# Each entry: canonical output name -> list of normalized aliases it may
# appear as in either input file.
CUSTOMER_INFO_COLUMNS: list[tuple[str, list[str]]] = [
    ("SAP ID", ["sap id", "new sap id"]),
    ("Customer Name", ["customer name"]),
    ("Customer Type", ["customer type"]),
    ("Region", ["region"]),
    ("Area", ["area"]),
    ("Territory", ["territory"]),
    ("TSM/TSE", ["tsm/tse", "tsm tse"]),
    ("ASM/KAM", ["asm/kam", "asm kam"]),
    ("RSM/B2B Head", ["rsm/b2b head", "rsm b2b head"]),
]

# Columns that are *required* for a file to be usable at all.
REQUIRED_CUSTOMER_COLUMNS = ["SAP ID", "Customer Name"]

# Product groups, in the exact output order requested, each with the four
# sub-columns (Target, MTD Sales, Yesterday Sales, Achievement %) and the
# normalized aliases that map onto each sub-column. Order of aliases matters
# only in that the first match found wins if a file somehow has duplicates.
PRODUCT_GROUPS: dict[str, dict[str, list[str]]] = {
    "PCC + OPC": {
        "Target": ["target (pcc + opc)", "target pcc + opc", "pcc + opc target", "pcc+opc target"],
        "MTD Sales": ["mtd pcc+opc sales", "mtd pcc + opc sales", "pcc+opc mtd sales"],
        "Yesterday Sales": ["yesterday pcc+opc sales", "yesterday pcc + opc sales"],
        "Achievement %": ["(pcc + opc) ach%", "pcc + opc ach%", "pcc+opc ach%"],
    },
    "HWP": {
        "Target": ["hwp target"],
        "MTD Sales": ["hwp mtd sales"],
        "Yesterday Sales": ["yesterday hwp sales"],
        "Achievement %": ["hwp ach%"],
    },
    "HCG": {
        "Target": ["hcg target"],
        "MTD Sales": ["hcg mtd sales"],
        "Yesterday Sales": ["yesterday hcg sales"],
        "Achievement %": ["hcg ach%"],
    },
    "PLC": {
        "Target": ["plc target"],
        "MTD Sales": ["plc mtd sales"],
        "Yesterday Sales": ["yesterday plc sales"],
        "Achievement %": ["plc ach%"],
    },
    "PLC+": {
        "Target": ["plc+ target"],
        "MTD Sales": ["plc+ mtd sales"],
        "Yesterday Sales": ["yesterday plc+ sales"],
        "Achievement %": ["plc+ ach%"],
    },
    "Powercrete": {
        "Target": ["powercrete target"],
        "MTD Sales": ["powercrete mtd sales"],
        # The source file has a standing typo ("Prowercrete") in this one
        # header only (Target/MTD Sales/Ach% are spelled correctly). Without
        # this alias the column silently fails to match and Yesterday Sales
        # for Powercrete comes back blank/0 in every merge, even though the
        # real figure is present in the file.
        "Yesterday Sales": ["yesterday powercrete sales", "yesterday prowercrete sales"],
        "Achievement %": ["powercrete ach%"],
    },
}

SUB_COLS = ["Target", "MTD Sales", "Yesterday Sales", "Achievement %"]

# Fallback aliases used only when the primary "MTD Sales" alias for PCC+OPC
# is missing, so a per-brand split (PCC MTD Sales + OPC MTD Sales) can be
# summed instead of losing the figure entirely.
PCC_OPC_MTD_FALLBACK_PARTS = ["pcc mtd sales", "opc mtd sales"]

# Fingerprint columns used purely to decide which uploaded file is which.
# A file is classified as "file1" (PLC/Powercrete) or "file2" (PCC+OPC/HWP/HCG)
# based on which fingerprint set it has more hits from.
FILE1_FINGERPRINT = ["plc target", "plc+ target", "powercrete target", "plc mtd sales"]
FILE2_FINGERPRINT = ["hwp target", "hcg target", "pcc + opc", "new sap id", "pcc mtd sales"]


# ---------------------------------------------------------------------------
# 2. Header normalization
# ---------------------------------------------------------------------------

def normalize_header(name: object) -> str:
    """Collapse whitespace/line breaks, strip, and lowercase a column name
    so headers that differ only in spacing/casing/line-wraps still match.
    """
    if name is None:
        return ""
    text = str(name)
    text = text.replace("\n", " ").replace("\r", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip().lower()


def build_normalized_lookup(columns) -> dict[str, str]:
    """Map normalized header -> original column name for a DataFrame's columns."""
    lookup: dict[str, str] = {}
    for col in columns:
        norm = normalize_header(col)
        if norm and norm not in lookup:  # first occurrence wins
            lookup[norm] = col
    return lookup


def find_column(lookup: dict[str, str], aliases: list[str]) -> Optional[str]:
    """Return the original column name matching any alias, or None."""
    for alias in aliases:
        if alias in lookup:
            return lookup[alias]
    return None


# ---------------------------------------------------------------------------
# 3. Loading + file-type detection
# ---------------------------------------------------------------------------

@dataclass
class LoadedFile:
    df: pd.DataFrame
    lookup: dict[str, str]
    label: str  # "file1" or "file2", for error messages only


ALL_KNOWN_ALIASES: list[str] = [alias for _canon, aliases in CUSTOMER_INFO_COLUMNS for alias in aliases] + [
    alias
    for subfields in PRODUCT_GROUPS.values()
    for aliases in subfields.values()
    for alias in aliases
]

MAX_HEADER_SCAN_ROWS = 25  # how many leading rows to inspect for the real header


def detect_header_row(raw: pd.DataFrame) -> int:
    """Real-world exports often have a title/summary block (working-day
    counters, a month banner, merged decorative cells, etc.) sitting above
    the actual column headers. Rather than assume row 0 is the header, scan
    the first few rows and pick whichever one has the most cells matching
    our known column aliases.
    """
    best_row, best_score = 0, -1
    scan_limit = min(MAX_HEADER_SCAN_ROWS, len(raw))
    for row_idx in range(scan_limit):
        row_values = [normalize_header(v) for v in raw.iloc[row_idx].tolist()]
        score = sum(1 for v in row_values if v in ALL_KNOWN_ALIASES)
        if score > best_score:
            best_row, best_score = row_idx, score

    if best_score <= 0:
        # Nothing matched anywhere in the scan window; fall back to row 0
        # so downstream validation can raise its usual clear error instead
        # of a confusing header-detection failure.
        return 0
    return best_row


def load_excel(path: str) -> pd.DataFrame:
    """Load the first sheet of an Excel file, auto-detecting which row
    holds the real column headers (skipping any title/summary rows above
    it), and return a strings-preserved DataFrame.
    """
    raw = pd.read_excel(path, header=None, dtype=object)
    header_row = detect_header_row(raw)

    new_columns = raw.iloc[header_row].tolist()
    df = raw.iloc[header_row + 1:].copy()
    df.columns = new_columns
    df.reset_index(drop=True, inplace=True)

    # Drop fully-blank rows/columns that sometimes trail a decorative block.
    df.dropna(axis=0, how="all", inplace=True)
    df.reset_index(drop=True, inplace=True)
    return df


def score_fingerprint(lookup: dict[str, str], fingerprint: list[str]) -> int:
    return sum(1 for alias in fingerprint if alias in lookup)


def classify_files(path_a: str, path_b: str) -> tuple[LoadedFile, LoadedFile]:
    """Load both files and decide which is the PLC/Powercrete file (file1)
    and which is the PCC+OPC/HWP/HCG file (file2), regardless of upload
    order or filename.
    """
    df_a, df_b = load_excel(path_a), load_excel(path_b)
    lookup_a, lookup_b = build_normalized_lookup(df_a.columns), build_normalized_lookup(df_b.columns)

    score_a1, score_a2 = score_fingerprint(lookup_a, FILE1_FINGERPRINT), score_fingerprint(lookup_a, FILE2_FINGERPRINT)
    score_b1, score_b2 = score_fingerprint(lookup_b, FILE1_FINGERPRINT), score_fingerprint(lookup_b, FILE2_FINGERPRINT)

    a_is_file1 = (score_a1 - score_a2) >= (score_b1 - score_b2)

    if a_is_file1:
        file1 = LoadedFile(df_a, lookup_a, f"'{path_a}' (PLC/Powercrete file)")
        file2 = LoadedFile(df_b, lookup_b, f"'{path_b}' (PCC+OPC/HWP/HCG file)")
    else:
        file1 = LoadedFile(df_b, lookup_b, f"'{path_b}' (PLC/Powercrete file)")
        file2 = LoadedFile(df_a, lookup_a, f"'{path_a}' (PCC+OPC/HWP/HCG file)")

    return file1, file2


# ---------------------------------------------------------------------------
# 4. Column validation
# ---------------------------------------------------------------------------

class MissingColumnsError(Exception):
    """Raised with a clear, itemized list of columns that could not be found."""


def validate_required_columns(file1: LoadedFile, file2: LoadedFile) -> None:
    missing: list[str] = []
    for canonical, aliases in CUSTOMER_INFO_COLUMNS:
        if canonical not in REQUIRED_CUSTOMER_COLUMNS:
            continue
        found_in_1 = find_column(file1.lookup, aliases)
        found_in_2 = find_column(file2.lookup, aliases)
        if not found_in_1 and not found_in_2:
            missing.append(f"{canonical} (expected one of: {', '.join(aliases)})")

    if missing:
        raise MissingColumnsError(
            "The following required columns were not found in either uploaded file:\n  - "
            + "\n  - ".join(missing)
        )


# ---------------------------------------------------------------------------
# 5. Extraction: pull customer-info + product data out of each raw file
# ---------------------------------------------------------------------------

def to_numeric(series: pd.Series) -> pd.Series:
    """Vectorized, safe numeric coercion. Handles blanks and stray text."""
    cleaned = series.astype(str).str.replace(",", "", regex=False).str.strip()
    cleaned = cleaned.replace({"": None, "nan": None, "None": None, "-": None})
    return pd.to_numeric(cleaned, errors="coerce")


def extract_customer_info(loaded: LoadedFile) -> pd.DataFrame:
    out = pd.DataFrame(index=loaded.df.index)
    for canonical, aliases in CUSTOMER_INFO_COLUMNS:
        col = find_column(loaded.lookup, aliases)
        out[canonical] = loaded.df[col] if col is not None else pd.NA
    return out


def extract_product_data(loaded: LoadedFile, groups: dict[str, dict[str, list[str]]]) -> pd.DataFrame:
    """Pull Target/MTD Sales/Yesterday Sales/Achievement % for whichever
    product groups actually have matching columns in this file.
    """
    out = pd.DataFrame(index=loaded.df.index)
    for group_name, subfields in groups.items():
        for sub_name, aliases in subfields.items():
            out_col = f"{group_name}||{sub_name}"
            col = find_column(loaded.lookup, aliases)

            if col is None and group_name == "PCC + OPC" and sub_name == "MTD Sales":
                # Fall back to summing the PCC/OPC per-brand MTD columns.
                part_cols = [find_column(loaded.lookup, [alias]) for alias in PCC_OPC_MTD_FALLBACK_PARTS]
                part_cols = [c for c in part_cols if c is not None]
                if part_cols:
                    out[out_col] = loaded.df[part_cols].apply(to_numeric).sum(axis=1, min_count=1)
                    continue

            if col is None:
                out[out_col] = pd.NA
                continue

            if sub_name == "Achievement %":
                out[out_col] = to_numeric(loaded.df[col])
            else:
                out[out_col] = to_numeric(loaded.df[col])
    return out


# ---------------------------------------------------------------------------
# 6. Merge
# ---------------------------------------------------------------------------

def coalesce(*series: pd.Series) -> pd.Series:
    result = series[0]
    for s in series[1:]:
        result = result.combine_first(s)
    return result


def normalize_join_name(series: pd.Series) -> pd.Series:
    """Normalized (trimmed, collapsed-whitespace, lowercased) Customer Name
    used only as part of the join key — never shown in output.
    """
    return series.astype(str).str.strip().str.lower().str.replace(r"\s+", " ", regex=True)


def normalize_territory(series: pd.Series) -> pd.Series:
    """Normalized Territory used only for the fallback-match comparison below."""
    return series.astype(str).str.strip().str.lower().str.replace(r"\s+", " ", regex=True)


def apply_territory_fallback_matches(side1: pd.DataFrame, side2: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Recovers matches lost when one file numbers a customer's sub-distributor
    rows ("Acme 1", "Acme 2", "Acme 3"...) while the other file keeps a single
    combined row for the same SAP ID ("Acme", no suffix). An exact match on
    (SAP ID, normalized Customer Name) misses these entirely -- the combined
    row never matches any of the numbered rows -- so its figures silently end
    up stranded in a disconnected output row instead of attached to the real
    sub-distributor(s).

    For each SAP ID where some rows fail to match by name on *both* sides,
    this falls back to (SAP ID, Territory): when the side with exactly one
    unmatched row shares its Territory with one or more unmatched rows on the
    other side, its figures are copied onto every one of those matching rows
    (the combined total is broadcast onto each matching sub-row, rather than
    split or dropped) by giving synthetic copies the sub-rows' `_join_name`
    so the normal outer join naturally attaches them.

    SAP IDs where neither side narrows down to exactly one unmatched row are
    left as-is (too ambiguous to guess safely) -- same as current behavior.
    """
    idx1 = pd.MultiIndex.from_frame(side1[["SAP ID", "_join_name"]])
    idx2 = pd.MultiIndex.from_frame(side2[["SAP ID", "_join_name"]])
    matched_keys = idx1.intersection(idx2)

    unmatched1 = side1[~idx1.isin(matched_keys)]
    unmatched2 = side2[~idx2.isin(matched_keys)]

    common_ids = set(unmatched1["SAP ID"]) & set(unmatched2["SAP ID"])
    if not common_ids:
        return side1, side2

    territory1 = normalize_territory(side1["Territory"])
    territory2 = normalize_territory(side2["Territory"])

    extra_side1_rows: list[pd.Series] = []
    extra_side2_rows: list[pd.Series] = []
    consumed_side1_idx: list = []
    consumed_side2_idx: list = []

    for sap_id in common_ids:
        u1 = unmatched1[unmatched1["SAP ID"] == sap_id]
        u2 = unmatched2[unmatched2["SAP ID"] == sap_id]

        # file2 holds the lone combined row; file1 holds the numbered sub-rows.
        if len(u2) == 1 and len(u1) >= 1:
            agg_row = u2.iloc[0]
            agg_territory = territory2.loc[agg_row.name]
            targets = u1[territory1.loc[u1.index] == agg_territory]
            if len(targets) >= 1:
                for sub_idx, sub in targets.iterrows():
                    copy = agg_row.copy()
                    copy["_join_name"] = sub["_join_name"]
                    extra_side2_rows.append(copy)
                consumed_side2_idx.append(agg_row.name)
                continue  # this SAP ID is resolved

        # file1 holds the lone combined row; file2 holds the numbered sub-rows.
        if len(u1) == 1 and len(u2) >= 1:
            agg_row = u1.iloc[0]
            agg_territory = territory1.loc[agg_row.name]
            targets = u2[territory2.loc[u2.index] == agg_territory]
            if len(targets) >= 1:
                for sub_idx, sub in targets.iterrows():
                    copy = agg_row.copy()
                    copy["_join_name"] = sub["_join_name"]
                    extra_side1_rows.append(copy)
                consumed_side1_idx.append(agg_row.name)

    if consumed_side2_idx:
        side2 = side2.drop(index=consumed_side2_idx)
    if consumed_side1_idx:
        side1 = side1.drop(index=consumed_side1_idx)

    if extra_side2_rows:
        side2 = pd.concat([side2, pd.DataFrame(extra_side2_rows)], ignore_index=True)
    if extra_side1_rows:
        side1 = pd.concat([side1, pd.DataFrame(extra_side1_rows)], ignore_index=True)

    return side1, side2


def build_merged_dataframe(file1: LoadedFile, file2: LoadedFile) -> pd.DataFrame:
    info1 = extract_customer_info(file1)
    info2 = extract_customer_info(file2)

    prod1 = extract_product_data(file1, {k: v for k, v in PRODUCT_GROUPS.items() if k in ("PLC", "PLC+", "Powercrete")})
    prod2 = extract_product_data(file2, {k: v for k, v in PRODUCT_GROUPS.items() if k in ("PCC + OPC", "HWP", "HCG")})

    # Normalize SAP ID to string for a reliable join key (avoids float vs. str mismatches).
    info1["SAP ID"] = info1["SAP ID"].astype(str).str.strip()
    info2["SAP ID"] = info2["SAP ID"].astype(str).str.strip()

    # One SAP ID commonly covers several sub-distributor/territory rows (e.g.
    # "J.I.Impex Ltd 1" .. "J.I.Impex Ltd 6" all under one SAP ID). Joining on
    # SAP ID alone would cross-join every such row in file1 against every one
    # in file2 for that ID, fabricating combinations that don't exist. Adding
    # normalized Customer Name to the join key pairs each sub-row with its
    # real counterpart instead.
    info1["_join_name"] = normalize_join_name(info1["Customer Name"])
    info2["_join_name"] = normalize_join_name(info2["Customer Name"])

    side1 = pd.concat([info1, prod1], axis=1)
    side2 = pd.concat([info2, prod2], axis=1)

    # Recover rows that would otherwise be lost/stranded because one file
    # numbers sub-distributor rows while the other keeps a single combined
    # row for the same SAP ID (see docstring above).
    side1, side2 = apply_territory_fallback_matches(side1, side2)

    merged = pd.merge(
        side1, side2,
        on=["SAP ID", "_join_name"], how="outer", suffixes=("_f1", "_f2"),
    )
    merged.drop(columns=["_join_name"], inplace=True)

    # Reconcile duplicated customer-info columns (present in both files) by
    # coalescing: prefer file1's value, fall back to file2's.
    for canonical, _aliases in CUSTOMER_INFO_COLUMNS:
        if canonical == "SAP ID":
            continue
        col_f1, col_f2 = f"{canonical}_f1", f"{canonical}_f2"
        if col_f1 in merged.columns and col_f2 in merged.columns:
            merged[canonical] = coalesce(merged[col_f1], merged[col_f2])
            merged.drop(columns=[col_f1, col_f2], inplace=True)
        elif col_f1 in merged.columns:
            merged.rename(columns={col_f1: canonical}, inplace=True)
        elif col_f2 in merged.columns:
            merged.rename(columns={col_f2: canonical}, inplace=True)

    return merged


# ---------------------------------------------------------------------------
# 7. Output-column assembly (final ordered layout)
# ---------------------------------------------------------------------------

def assemble_output_frame(merged: pd.DataFrame) -> pd.DataFrame:
    ordered_cols: list[str] = [c for c, _ in CUSTOMER_INFO_COLUMNS]
    header_top: list[str] = ["Customer Information"] * len(ordered_cols)
    header_sub: list[str] = list(ordered_cols)

    data = {c: merged[c] for c in ordered_cols}

    for group_name in PRODUCT_GROUPS:
        for sub_name in SUB_COLS:
            key = f"{group_name}||{sub_name}"
            flat_name = f"{group_name} - {sub_name}"
            ordered_cols.append(flat_name)
            header_top.append(group_name)
            header_sub.append(sub_name)
            data[flat_name] = merged[key] if key in merged.columns else pd.Series([pd.NA] * len(merged))

    out = pd.DataFrame(data)[ordered_cols]
    out.attrs["header_top"] = header_top
    out.attrs["header_sub"] = header_sub
    return out


# ---------------------------------------------------------------------------
# 8. Excel export with professional formatting
# ---------------------------------------------------------------------------

THIN_BORDER = Border(*[Side(style="thin", color="B7B7B7")] * 4)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
GROUP_FILL = {
    "Customer Information": PatternFill("solid", fgColor="404040"),
    "PCC + OPC": PatternFill("solid", fgColor="1F4E78"),
    "HWP": PatternFill("solid", fgColor="2E5F8A"),
    "HCG": PatternFill("solid", fgColor="3D7AB5"),
    "PLC": PatternFill("solid", fgColor="4F81BD"),
    "PLC+": PatternFill("solid", fgColor="6699CC"),
    "Powercrete": PatternFill("solid", fgColor="8DB4E2"),
}
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
SUBHEADER_FONT = Font(bold=True, color="FFFFFF", size=9)
BODY_FONT = Font(size=9)
ZEBRA_FILL = PatternFill("solid", fgColor="F2F2F2")

GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
GREEN_FONT = Font(color="006100")
YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")
YELLOW_FONT = Font(color="9C6500")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
RED_FONT = Font(color="9C0006")


def export_workbook(out: pd.DataFrame, output_path: str) -> None:
    header_top = out.attrs["header_top"]
    header_sub = out.attrs["header_sub"]
    n_customer_cols = len(CUSTOMER_INFO_COLUMNS)
    n_cols = len(out.columns)
    n_rows = len(out)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Summary"

    # --- Header rows (row 1 = group title, row 2 = sub-column title) ---
    # Customer-info columns have no sub-group: their real column name (SAP ID,
    # Customer Name, ...) is written into row 1 (the cell that survives the
    # vertical merge below) so it isn't overwritten by a generic group label.
    for col_idx in range(1, n_cols + 1):
        is_customer_info = col_idx <= n_customer_cols
        row1_value = header_sub[col_idx - 1] if is_customer_info else header_top[col_idx - 1]
        row2_value = None if is_customer_info else header_sub[col_idx - 1]

        c1 = ws.cell(row=1, column=col_idx, value=row1_value)
        c2 = ws.cell(row=2, column=col_idx, value=row2_value)
        fill = GROUP_FILL.get(header_top[col_idx - 1], HEADER_FILL)
        for cell, font in ((c1, HEADER_FONT), (c2, SUBHEADER_FONT)):
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER

    # Merge customer-info headers vertically (row1:row2) since they have no sub-groups.
    for col_idx in range(1, n_customer_cols + 1):
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)

    # Merge each product group's header horizontally across its 4 sub-columns.
    col_idx = n_customer_cols + 1
    for _group_name in PRODUCT_GROUPS:
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx + len(SUB_COLS) - 1)
        col_idx += len(SUB_COLS)

    # --- Data rows ---
    achievement_cols: list[int] = []
    for col_idx, col_name in enumerate(out.columns, start=1):
        if header_sub[col_idx - 1] == "Achievement %":
            achievement_cols.append(col_idx)

    for row_offset, (_, row) in enumerate(out.iterrows()):
        excel_row = row_offset + 3
        is_zebra = row_offset % 2 == 1
        for col_idx, col_name in enumerate(out.columns, start=1):
            value = row[col_name]
            if pd.isna(value):
                value = None
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            if is_zebra:
                cell.fill = ZEBRA_FILL

            sub_label = header_sub[col_idx - 1]
            if sub_label == "Achievement %":
                cell.number_format = "0.00%"
                cell.alignment = Alignment(horizontal="center")
            elif sub_label in ("Target", "MTD Sales", "Yesterday Sales"):
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")

    # --- Conditional formatting on every Achievement % column ---
    if n_rows > 0:
        for col_idx in achievement_cols:
            col_letter = get_column_letter(col_idx)
            rng = f"{col_letter}3:{col_letter}{n_rows + 2}"
            # Values are stored as fractions (0.85 == 85%), matching the 0.00% format.
            ws.conditional_formatting.add(
                rng, CellIsRule(operator="greaterThan", formula=["1"], fill=GREEN_FILL, font=GREEN_FONT)
            )
            ws.conditional_formatting.add(
                rng, CellIsRule(operator="between", formula=["0.8", "1"], fill=YELLOW_FILL, font=YELLOW_FONT)
            )
            ws.conditional_formatting.add(
                rng, CellIsRule(operator="lessThan", formula=["0.8"], fill=RED_FILL, font=RED_FONT)
            )

    # --- Column widths ---
    for col_idx, col_name in enumerate(out.columns, start=1):
        header_len = max(len(str(header_top[col_idx - 1])), len(str(header_sub[col_idx - 1])))
        try:
            max_data_len = out[col_name].astype(str).map(len).max()
        except Exception:
            max_data_len = 10
        width = max(header_len, int(max_data_len) if pd.notna(max_data_len) else 10, 10) + 2
        ws.column_dimensions[get_column_letter(col_idx)].width = min(width, 28)

    # --- Freeze panes: freeze header rows + customer-info columns ---
    freeze_col_letter = get_column_letter(n_customer_cols + 1)
    ws.freeze_panes = f"{freeze_col_letter}3"

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 22

    wb.save(output_path)


# ---------------------------------------------------------------------------
# 9. Orchestration
# ---------------------------------------------------------------------------

def run(path_a: str, path_b: str, output_path: str = "Sales_Summary.xlsx") -> str:
    file1, file2 = classify_files(path_a, path_b)
    validate_required_columns(file1, file2)
    merged = build_merged_dataframe(file1, file2)

    # Recompute Achievement % as MTD Sales / Target wherever the source value
    # is missing but both inputs are present — keeps the figure trustworthy
    # even if the source Ach% column was blank.
    for group_name in PRODUCT_GROUPS:
        target_key = f"{group_name}||Target"
        mtd_key = f"{group_name}||MTD Sales"
        ach_key = f"{group_name}||Achievement %"
        if target_key in merged.columns and mtd_key in merged.columns and ach_key in merged.columns:
            computed = merged[mtd_key] / merged[target_key].replace({0: pd.NA})
            merged[ach_key] = merged[ach_key].combine_first(computed)

    out = assemble_output_frame(merged)
    # Sort by SAP ID for stable, predictable output.
    out = out.sort_values("SAP ID", kind="stable").reset_index(drop=True)
    out.attrs["header_top"], out.attrs["header_sub"] = (
        ["Customer Information"] * len(CUSTOMER_INFO_COLUMNS)
        + [g for g in PRODUCT_GROUPS for _ in SUB_COLS],
        [c for c, _ in CUSTOMER_INFO_COLUMNS] + [s for _ in PRODUCT_GROUPS for s in SUB_COLS],
    )

    export_workbook(out, output_path)
    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(description="Merge Supercrete + PCC/OPC distribution files into Sales_Summary.xlsx")
    parser.add_argument("file_a", help="Path to the first uploaded Excel file")
    parser.add_argument("file_b", help="Path to the second uploaded Excel file")
    parser.add_argument("-o", "--output", default="Sales_Summary.xlsx", help="Output workbook path")
    args = parser.parse_args()

    try:
        output_path = run(args.file_a, args.file_b, args.output)
    except MissingColumnsError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        sys.exit(1)

    print(output_path)


if __name__ == "__main__":
    main()